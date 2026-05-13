import argparse
import importlib.util
import json
import os
import re
import sys
from pathlib import Path
from typing import Dict, Iterable, List, Optional

import torch
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PIL import Image
from torchvision import transforms

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from config import BASE_WORKSPACE


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_DLKGS_DIR = PROJECT_ROOT / "DLKGS"
DEFAULT_MODEL_FILE = DEFAULT_DLKGS_DIR / "five_model" / "model.py"
DEFAULT_WEIGHTS_DIR = DEFAULT_DLKGS_DIR / "five_model" / "model_weights"
DEFAULT_OUTPUT_ROOT = PROJECT_ROOT / "result" / "DLKGS"

LABEL_ORDER = ["A", "DR", "H", "IM", "N"]
LABEL_NUM_CLASSES = {
    "A": 4,
    "DR": 4,
    "H": 3,
    "IM": 4,
    "N": 3,
}
WEIGHT_FILES = {
    "A": "model_0.pth",
    "DR": "model_1.pth",
    "H": "model_2.pth",
    "IM": "model_3.pth",
    "N": "model_4.pth",
}
CLASS_NAME_CONFIG = {
    "A": ["NO", "A0", "A1", "A2"],
    "DR": ["NO", "DR0", "DR1", "DR2"],
    "H": ["NO", "H0", "H1"],
    "IM": ["NO", "IM0", "IM1", "IM2"],
    "N": ["NO", "N0", "N1"],
}
DATASET_ALIASES = {
    "batch": "testdataset",
    "test": "testdataset",
    "testdataset": "testdataset",
    "external": "externaldataset",
    "externaldataset": "externaldataset",
}
IMAGE_SUFFIXES = (".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff")
PREDICTION_HEADER = [
    "filepath",
    "predicted_labels",
    "predicted_scores",
    "score",
    "groundtruth",
    "is_correct",
]


def resolve_dataset_name(dataset: str) -> str:
    try:
        return DATASET_ALIASES[dataset.lower()]
    except KeyError as exc:
        choices = ", ".join(sorted(DATASET_ALIASES))
        raise ValueError(f"Unknown dataset '{dataset}'. Choose one of: {choices}") from exc


def default_base_dir(dataset_name: str) -> Path:
    return Path(BASE_WORKSPACE) / dataset_name


def default_json_output(dataset_name: str) -> Path:
    return DEFAULT_OUTPUT_ROOT / dataset_name / "test_results.json"


def default_excel_output(dataset_name: str) -> Path:
    return DEFAULT_OUTPUT_ROOT / dataset_name / "DLKGs_predictions.xlsx"


def load_five_model_module(model_file: Path):
    if not model_file.exists():
        raise FileNotFoundError(f"DLKGS model file does not exist: {model_file}")

    spec = importlib.util.spec_from_file_location("dlkgs_five_model_module", model_file)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module


def load_models(model_file: Path, weights_dir: Path, device: torch.device) -> Dict[str, torch.nn.Module]:
    module = load_five_model_module(model_file)
    create_model = module.efficientnetv2_l

    models = {}
    for label, num_classes in LABEL_NUM_CLASSES.items():
        weight_path = weights_dir / WEIGHT_FILES[label]
        if not weight_path.exists():
            raise FileNotFoundError(f"DLKGS weights file does not exist: {weight_path}")

        model = create_model(num_classes=num_classes).to(device)
        state_dict = torch.load(weight_path, map_location=device)
        model.load_state_dict(state_dict)
        model.eval()
        models[label] = model
        print(f"loaded {label} from {weight_path}")

    return models


def build_transform():
    return transforms.Compose(
        [
            transforms.Resize(480),
            transforms.CenterCrop(480),
            transforms.ToTensor(),
            transforms.Normalize([0.5, 0.5, 0.5], [0.5, 0.5, 0.5]),
        ]
    )


def iter_image_files(base_dir: Path) -> Iterable[Path]:
    for category in LABEL_ORDER:
        category_dir = base_dir / category
        if not category_dir.exists():
            print(f"skip missing directory: {category_dir}")
            continue

        for sub_dir in sorted(item for item in category_dir.iterdir() if item.is_dir()):
            image_files = sorted(item for item in sub_dir.iterdir() if item.suffix.lower() in IMAGE_SUFFIXES)
            print(f"processing {category}/{sub_dir.name}, images: {len(image_files)}")
            yield from image_files


def predict_image(models, image_path: Path, transform, device: torch.device):
    image = Image.open(image_path).convert("RGB")
    image_tensor = transform(image).unsqueeze(0).to(device)

    predictions = {}
    with torch.no_grad():
        for label, model in models.items():
            output = model(image_tensor)
            pred_idx = int(torch.argmax(output, dim=1).item())
            scores = torch.softmax(output, dim=1).squeeze(0).detach().cpu().tolist()
            predictions[label] = {
                "predicted_label": pred_idx,
                "scores": [float(score) for score in scores],
            }
    return predictions


def collect_results(base_dir: Path, models, transform, device: torch.device):
    results = []
    print(f"start processing: {base_dir}")

    for image_file in iter_image_files(base_dir):
        try:
            sub_dir = image_file.parent
            category_dir = sub_dir.parent
            results.append(
                {
                    "category": category_dir.name,
                    "sub_category": sub_dir.name,
                    "filename": image_file.name,
                    "filepath": str(image_file),
                    "predictions": predict_image(models, image_file, transform, device),
                }
            )
        except Exception as exc:
            print(f"failed on {image_file}: {exc}")

    return results


def save_results_json(results, weights_dir: Path, output_path: Path):
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as file:
        json.dump(
            {
                "weights_dir": str(weights_dir),
                "results": results,
            },
            file,
            ensure_ascii=False,
            indent=2,
        )


def load_results(json_path: Path):
    with open(json_path, "r", encoding="utf-8") as file:
        payload = json.load(file)

    if isinstance(payload, dict) and "results" in payload:
        return payload["results"]
    return payload


def get_true_label(sub_category: str) -> int:
    sub = sub_category.upper()
    if sub == "NO":
        return 0

    match = re.search(r"\d+", sub)
    if match:
        return int(match.group()) + 1

    return -1


def label_to_subclass_name(category: str, pred_label: Optional[int]):
    names = CLASS_NAME_CONFIG.get(category, [])
    if pred_label is None:
        return None

    pred_label = int(pred_label)
    if 0 <= pred_label < len(names):
        return names[pred_label]
    return None


def to_relative_filepath(filepath: str, dataset_root: Path) -> str:
    path = Path(filepath)
    try:
        return str(path.relative_to(dataset_root))
    except ValueError:
        filepath_str = str(path)
        marker = str(dataset_root) + os.sep
        if filepath_str.startswith(marker):
            return filepath_str[len(marker) :]
        return filepath_str


def build_excel_rows(results, dataset_root: Path):
    rows = []
    for item in results:
        filepath = to_relative_filepath(item.get("filepath", ""), dataset_root)
        category = item.get("category", "")
        sub_category = item.get("sub_category", "")
        predictions = item.get("predictions", {})

        predicted_labels = []
        predicted_scores = []
        for label in LABEL_ORDER:
            pred_info = predictions.get(label, {})
            pred_label = pred_info.get("predicted_label")
            scores = pred_info.get("scores", [])

            if pred_label is None:
                predicted_labels.append(None)
                predicted_scores.append(None)
                continue

            pred_label = int(pred_label)
            predicted_labels.append(pred_label)
            if 0 <= pred_label < len(scores):
                predicted_scores.append(float(scores[pred_label]))
            else:
                predicted_scores.append(None)

        current_category_prediction = predictions.get(category, {}).get("predicted_label")
        true_label = get_true_label(sub_category)
        is_correct = (
            int(current_category_prediction == true_label)
            if true_label >= 0 and current_category_prediction is not None
            else None
        )
        score = label_to_subclass_name(category, current_category_prediction)

        rows.append(
            [
                filepath,
                json.dumps(predicted_labels, ensure_ascii=False),
                json.dumps(predicted_scores, ensure_ascii=False),
                score,
                sub_category,
                is_correct,
            ]
        )

    return rows


def style_predictions_sheet(ws):
    no_fill = PatternFill(fill_type=None)
    thin_side = Side(style="thin", color="000000")
    header_font = Font(name="Consolas", size=11, bold=True)
    header_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    header_alignment = Alignment(horizontal="center", vertical="top")
    text_alignment = Alignment(horizontal="left", vertical="center")
    center_alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.font = header_font
        cell.border = header_border
        cell.alignment = header_alignment
        cell.fill = no_fill

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.fill = no_fill
            cell.alignment = text_alignment if cell.column <= 3 else center_alignment

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 70
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 12
    ws.freeze_panes = "A2"


def add_note_sheet(wb):
    note = wb.create_sheet("note")
    note.append(["field", "meaning"])
    note.append(["filepath", r"Relative path under dataset root, e.g. A\A0\1.jpg"])
    note.append(["predicted_labels", "Prediction list in order [A, DR, H, IM, N], from each predicted_label"])
    note.append(["predicted_scores", "Score list in order [A, DR, H, IM, N], each value is scores[predicted_label]"])
    note.append(["score", "Predicted subclass name for the sample's own category, e.g. A1, DR2, H0, NO"])
    note.append(["groundtruth", "True subclass label parsed from filepath, e.g. A0, DR1, IM2, NO"])
    note.append(["is_correct", "1 if the prediction of the sample's own category equals the true label, else 0"])
    note.column_dimensions["A"].width = 18
    note.column_dimensions["B"].width = 100

    for cell in note[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type=None)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def save_excel(results, dataset_root: Path, output_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "predictions"
    ws.append(PREDICTION_HEADER)

    for row in build_excel_rows(results, dataset_root):
        ws.append(row)

    style_predictions_sheet(ws)
    add_note_sheet(wb)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def parse_args():
    parser = argparse.ArgumentParser(description="Run DLKGS five-model prediction and export Excel results.")
    parser.add_argument(
        "--dataset",
        type=str,
        default="testdataset",
        choices=sorted(DATASET_ALIASES),
        help="Dataset split alias. Use testdataset/batch for batch test set or externaldataset/external for external test set.",
    )
    parser.add_argument(
        "--base-dir",
        type=str,
        default=None,
        help="Override dataset root containing A/DR/H/IM/N subdirectories.",
    )
    parser.add_argument(
        "--model-file",
        type=str,
        default=str(DEFAULT_MODEL_FILE),
        help="Path to DLKGS five_model/model.py.",
    )
    parser.add_argument(
        "--weights-dir",
        type=str,
        default=str(DEFAULT_WEIGHTS_DIR),
        help="Directory containing model_0.pth to model_4.pth.",
    )
    parser.add_argument(
        "--json-output",
        type=str,
        default=None,
        help="Path to save intermediate test_results.json.",
    )
    parser.add_argument(
        "--excel-output",
        type=str,
        default=None,
        help="Path to save DLKGs_predictions.xlsx.",
    )
    parser.add_argument(
        "--device",
        type=str,
        default="cuda" if torch.cuda.is_available() else "cpu",
        help="Inference device, e.g. cuda or cpu.",
    )
    return parser.parse_args()


def main():
    args = parse_args()
    dataset_name = resolve_dataset_name(args.dataset)
    base_dir = Path(args.base_dir) if args.base_dir else default_base_dir(dataset_name)
    weights_dir = Path(args.weights_dir)
    model_file = Path(args.model_file)
    json_output = Path(args.json_output) if args.json_output else default_json_output(dataset_name)
    excel_output = Path(args.excel_output) if args.excel_output else default_excel_output(dataset_name)

    if not base_dir.exists():
        raise FileNotFoundError(f"dataset directory does not exist: {base_dir}")

    device = torch.device(args.device)
    models = load_models(model_file, weights_dir, device)
    transform = build_transform()
    results = collect_results(base_dir, models, transform, device)

    save_results_json(results, weights_dir, json_output)
    save_excel(load_results(json_output), base_dir, excel_output)

    print("--- DLKGS prediction finished ---")
    print(f"dataset: {dataset_name}")
    print(f"total images: {len(results)}")
    print(f"json output: {json_output}")
    print(f"excel output: {excel_output}")


if __name__ == "__main__":
    main()
